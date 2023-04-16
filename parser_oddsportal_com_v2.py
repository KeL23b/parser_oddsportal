import re
import time
import datetime
import os
import pickle
from multiprocessing import Process
from multiprocessing import freeze_support
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook
from time import sleep
from bs4 import BeautifulSoup


def login_acc(driver):
    """
    Регистрируется на сайте и создаёт файл с cookies
     или подгружает cookie из файла.
    :return: None
    """
    data_log_pass = open("login.txt", encoding='utf-8').readlines()
    login_in_acc = data_log_pass[0].strip()
    password_in_acc = data_log_pass[1].strip()
    driver.get("https://www.oddsportal.com")
    if not f"{login_in_acc}_cookies" in os.listdir(os.getcwd()):  # если нет файла с куками то он создаст его
        driver.get("https://www.oddsportal.com/login")
        driver.implicitly_wait(7)
        driver.find_element(By.ID, 'login-username-sign').send_keys(login_in_acc)
        driver.find_element(By.ID, 'login-password-sign').send_keys(password_in_acc)
        driver.execute_script('window.scrollBy(0, 500)')
        sleep(1)
        driver.find_element(By.NAME, 'login-submit').click()
        sleep(6)
        pickle.dump(driver.get_cookies(), open(f"{login_in_acc}_cookies", "wb"))
    else:
        for cookie in pickle.load(open(f"{login_in_acc}_cookies", "rb")):
            driver.add_cookie(cookie)
        sleep(2)
        driver.refresh()
        sleep(0.5)


def open_xlsx_file():
    """
    В директории с файлом ищет все .xlsx и создает на них Workbook object
    :return:[{'file_name': имя файла .xlsx, 'workbook': Workbook object}]
    """
    workbooks = []
    files_name = os.listdir(os.getcwd())
    for file_name in files_name:
        if '.xlsx' in file_name:
            workbook = load_workbook(file_name)
            workbooks.append(
                {
                    'file_name': file_name,
                    'workbook': workbook
                }
            )
    return workbooks


def get_links_to_parse(sheet):
    """
    :param sheet: лист из .xlsx файла с именем CONTROL
    :return: Список со ссылками(помеченными галочкой в .xlsx файле)
    """
    rows = sheet.max_row
    links_in_sheet = []
    for i in range(2, rows + 1):
        link = sheet.cell(row=i, column=1).value
        flag = sheet.cell(row=i, column=3).value
        if flag is not None:
            links_in_sheet.append(link)
    return links_in_sheet


def get_links_to_matches(links_matches: list, driver):
    """
    :param links_matches: Список с ссылками на матчи
    :param driver: driver: Объект selenium
    :return links_matches: Добавляет матчи и возвращает изменённый список
    """
    if driver.find_elements(By.CSS_SELECTOR, 'div.flex.flex-col.border-b.border-black-borders'):
        matches_selenium = driver.find_elements(By.CSS_SELECTOR, 'div.flex.flex-col.border-b.border-black-borders')
        for match in matches_selenium:
            if match.find_elements(By.TAG_NAME, 'a'):
                link_match = match.find_element(By.TAG_NAME, 'a').get_attribute('href')
                links_matches.append(link_match)
    return links_matches


def parses_links_to_matches(driver, workbook_dick, link_to_team):
    """
    Функция собирает ссылки на игры
    :param driver:
    :param workbook_dick:
    :param link_to_team: ссылка на команды
    :return: links_to_matches - список со ссылками на матчи, которых ещё нет в файле
             name_team - имя команды
    """
    links_to_matches = []

    driver.get(link_to_team)
    driver.implicitly_wait(3)
    time.sleep(2)
    driver.execute_script('window.scrollBy(0, document.body.scrollHeight);')
    driver.implicitly_wait(7)
    time.sleep(4)

    name_team = driver.find_element(By.XPATH, '//*[@id="search-match"]').get_attribute('value')
    create_new_sheet_if_not(workbook_dick, name_team)
    print(f'The command is being parsed: {name_team} link: {link_to_team}')
    links_to_matches = get_links_to_matches(links_to_matches, driver)

    """Получаем погиницию и проверяем записаны ли матчи в файле"""
    if workbook_dick['workbook'][name_team].cell(row=workbook_dick['workbook'][name_team].max_row,
                                                 column=26).value in links_to_matches:
        pages = 1
    elif driver.find_elements(By.CSS_SELECTOR, 'a.w-6.h-6.bg-no-repeat.bg-skip-next'):
        pages = driver.find_elements(By.CSS_SELECTOR, 'a.w-6.h-6.bg-no-repeat.bg-skip-next')
        pages = int(pages[1].get_attribute('href').split('/')[-2])
    else:
        pages = 1
    """ Проходи по погинации"""
    for page in range(2, pages + 1):
        driver.get(f'{link_to_team}page/{page}/')
        driver.implicitly_wait(3)
        time.sleep(4)
        driver.execute_script('window.scrollBy(0, document.body.scrollHeight);')
        driver.implicitly_wait(7)
        time.sleep(4)
        links_to_matches = get_links_to_matches(links_to_matches, driver)
        if workbook_dick['workbook'][name_team].cell(row=workbook_dick['workbook'][name_team].max_row,
                                                     column=26).value in links_to_matches:
            break
        print(f'{page} pages {name_team}')

    """ Удаляем из links_to_matches ссылки которые уже есть в файле"""
    if name_team in workbook_dick['workbook'].sheetnames:
        for i in range(2, workbook_dick['workbook'][name_team].max_row + 1):
            if workbook_dick['workbook'][name_team].cell(row=i, column=26).value in links_to_matches:
                links_to_matches.remove(workbook_dick['workbook'][name_team].cell(row=i, column=26).value)
    return links_to_matches, name_team


def parses_to_match(driver, link, name_team):
    """
    Парсит нужную информацию со страницы матча
    :param driver:
    :param link: Ссылка на игру
    :param name_team: Имя команды
    :return: Словарь с данными о матче
    """
    user_predictions = []
    abandoned = False
    over_time_penalties = False
    over_time = False
    error_goals = False
    postponed = False
    no_score = False
    match_is_today = False
    tabs_1_x_2 = False
    coefficient_is_not = False

    try:
        driver.get(link)
    except TimeoutException:
        time.sleep(5)
        driver.get(link)

    driver.implicitly_wait(7)
    driver.execute_script('window.scrollBy(0, document.body.scrollHeight);')
    time.sleep(1)
    soup_main = BeautifulSoup(driver.page_source, 'lxml')
    sport = driver.find_elements(By.CLASS_NAME, 'breadcrumb-link')[1].text
    country = driver.find_elements(By.CLASS_NAME, 'breadcrumb-link')[2].text
    liga = driver.find_elements(By.CLASS_NAME, 'breadcrumb-link')[3].text
    teams_not_formatted = driver.find_element(By.CLASS_NAME, 'capitalize.font-normal > p').text
    date_match_no_formatted = driver.find_element(By.CSS_SELECTOR,
                                                  'div.flex.text-xs.font-normal.text-gray-dark.font-main.item-center').text
    date_match = date_match_no_formatted.split(',')
    if driver.find_elements(By.CSS_SELECTOR, 'div.flex.flex-wrap > strong'):
        goals = driver.find_element(By.CSS_SELECTOR, 'div.flex.flex-wrap > strong').text
        account_not_formatted = driver.find_elements(By.CSS_SELECTOR, 'div.flex.flex-wrap')[2]\
            .text.replace('\n', '').strip()
    else:
        goals = 'None'
        account_not_formatted = None
        no_score = True
    first_command = \
        driver.find_elements(By.CSS_SELECTOR, 'div.flex.items-center.gap-1 > p')[
            0].text.strip()
    second_command = \
        driver.find_elements(By.CSS_SELECTOR, 'div.flex.items-center.gap-1 > p')[
            1].text.strip()
    if name_team in first_command:
        first_command = re.sub('\(\w*\)', '', first_command).strip()
    else:
        second_command = re.sub('\(\w*\)', '', second_command).strip()
    if goals != 'None' and re.sub('[a-zA-Z:]*', '', goals).strip().isdigit():  # если голы есть
        if 'ET' in goals or 'penalties' in goals:
            ET = soup_main.find('div', class_='flex flex-wrap')
            ET.find('span').decompose()
            ET.find('strong').decompose()
            data_goals = ET.text.strip()
            data_goals = data_goals.replace('(', '').replace(')', '').strip()
            if len(data_goals.split(',')) > 2:
                if 'ET' in goals:
                    over_time = True
                else:
                    over_time_penalties = True
                first_goal = int(data_goals.split(',')[0].split(':')[0]) + int(
                    data_goals.split(',')[0].split(':')[1])
                second_goal = int(data_goals.split(',')[1].split(':')[0]) + int(
                    data_goals.split(',')[1].split(':')[1])
                if first_goal > second_goal:
                    while first_goal != second_goal:
                        first_goal -= 1
                elif first_goal < second_goal:
                    while first_goal != second_goal:
                        second_goal -= 1
                first_goal = str(first_goal)
                second_goal = str(second_goal)
            else:
                error_goals = True
                first_goal = 'None'
                second_goal = 'None'
                print('Произошла ошибка при считывании голов')
        elif re.search('abandoned', driver.page_source):
            first_goal = 'None'
            second_goal = 'None'
            abandoned = True
            print('Матч прерван')
        else:
            goals = goals.split(':')
            first_goal = goals[0].strip()
            second_goal = goals[1].strip()
    elif re.search(f'{first_command} awarded', goals):
        first_goal = '1'
        second_goal = '0'
        no_score = True
    elif re.search(f'{second_command} awarded', goals):
        first_goal = '0'
        second_goal = '1'
        no_score = True
    elif len(driver.find_elements(By.CSS_SELECTOR, 'div.flex.flex-wrap > strong')) > 0:
        first_goal = 'None'
        second_goal = 'None'
        if driver.find_element(By.CSS_SELECTOR,
                               'div.flex.flex-wrap > strong').text == 'postponed':
            postponed = True  # если матч прерван
    else:
        first_goal = 'None'
        second_goal = 'None'
    if first_command.lower().replace(' ', '') in name_team.lower().replace(' ', ''):
        first_command = name_team
    else:
        second_command = name_team
    weekend = date_match[0]
    if 'Today' in weekend:
        match_is_today = True
    day = int(date_match[1].split()[0])
    mouth = date_match[1].split()[1]
    years = int(date_match[1].split()[2])
    hours = int(date_match[2].split(':')[0].strip())
    minutes = int(date_match[2].split(':')[1].strip())
    """Получаем коэффициенты """
    _1_x_2 = [tag.text for tag in driver.find_elements(By.CSS_SELECTOR, 'span.flex')]
    if '1X2' not in _1_x_2:
        tabs_1_x_2 = True
        coefficient_1 = ''
        coefficient_2 = ''
        coefficient_x = ''
    elif soup_main.find_all('div', class_='flex text-xs border-b h-9'):
        coefficients = soup_main.find_all(class_='flex text-xs border-b h-9')
        coefficient_1 = []
        coefficient_x = []
        coefficient_2 = []
        for coefficient in coefficients:
            if len(coefficient.find_all('p', class_='height-content')) <= 1:
                continue
            coefficient_1_2_x = coefficient.find_all('p', class_='height-content')
            try:
                coefficient_1.append(float(coefficient_1_2_x[1].text))
            except ValueError:
                pass
            try:
                coefficient_x.append(float(coefficient_1_2_x[2].text))
            except ValueError:
                pass
            try:
                coefficient_2.append(float(coefficient_1_2_x[3].text))
            except ValueError:
                pass
        if coefficient_1 and coefficient_x and coefficient_2:
            coefficient_1 = round(sum(coefficient_1) / len(coefficient_1), 2)
            coefficient_x = round(sum(coefficient_x) / len(coefficient_x), 2)
            coefficient_2 = round(sum(coefficient_2) / len(coefficient_2), 2)
        else:
            coefficient_is_not = True
            coefficient_1 = ''
            coefficient_2 = ''
            coefficient_x = ''
    else:
        coefficient_is_not = True
        coefficient_1 = ''
        coefficient_2 = ''
        coefficient_x = ''
    """Собираем количество проголосовавших"""
    if driver.find_elements(By.CSS_SELECTOR, 'div.cursor-pointer.underline') and not tabs_1_x_2:
        details = driver.find_element(By.CSS_SELECTOR, 'div.cursor-pointer.underline')
        driver.execute_script("arguments[0].scrollIntoView();", details)
        time.sleep(1)
        details.click()
        driver.implicitly_wait(3)
        time.sleep(1)
        user_predictions_tabs = driver.find_elements(By.CSS_SELECTOR,
                                                     'li.flex.items-center.border-b-2.cursor-pointer.border-black-main')
        driver.execute_script("arguments[0].scrollIntoView();", user_predictions_tabs[0])
        time.sleep(0.5)
        amount_people_result = driver.find_element(By.CSS_SELECTOR,
                                                   'p.flex.text-xs.font-normal.font-main.text-gray-dark').text
        amount_people_result = int(amount_people_result.split()[3])
        for tabs in user_predictions_tabs:
            tabs.click()
            time.sleep(0.5)
            driver.implicitly_wait(3)
            amount_people = driver.find_element(By.CSS_SELECTOR,
                                                'p.flex.text-xs.font-normal.font-main.text-gray-dark').text
            user_predictions.append(int(amount_people.split()[0]))
    else:
        amount_people_result = 0
        user_predictions = [0, 0, 0]

    if postponed:
        comment = 'Матч перенесён.'
    elif abandoned:
        comment = 'Матч прерван.'
    elif tabs_1_x_2:
        comment = 'Нет коэффициентов для события 1Х2'
    elif coefficient_is_not:
        comment = 'Нет одного или нескольких коэффициентов.'
    elif over_time:
        comment = 'Овертайм.'
    elif over_time_penalties:
        comment = 'Овертайм. Пенальти.'
    elif error_goals:
        comment = 'Произошла ошибка при считывании голов.'
    elif match_is_today:
        comment = 'Матч начался.'
    elif no_score:
        comment = 'Счёт неизвестен.'
    else:
        comment = None
    if first_goal == 'None' or second_goal == 'None':
        first_goal = None
        second_goal = None
    else:
        first_goal = int(first_goal)
        second_goal = int(second_goal)
    return {
        'name_team': name_team,
        'sport': sport,
        'country': country,
        'liga': liga,
        'date_match_no_formatted': date_match_no_formatted,
        'day': day,
        'mouth': mouth,
        'years': years,
        'weekend': weekend,
        'hours': hours,
        'minutes': minutes,
        'teams_not_formatted': teams_not_formatted,
        'first_command': first_command,
        'second_command': second_command,
        'account_not_formatted': account_not_formatted,
        'first_goal': first_goal,
        'second_goal': second_goal,
        'coefficient_1': coefficient_1,
        'coefficient_x': coefficient_x,
        'coefficient_2': coefficient_2,
        'user_predictions': user_predictions,
        'amount_people_result': amount_people_result,
        'comment': comment,
    }


def create_new_sheet_if_not(workbook_dick, name_team):
    """
    Проверяем наличие листа с именем команды в .xlsx, если его нет то создаём
    :param workbook_dick:
    :param name_team: Имя команды
    :return: None
    """
    if name_team not in workbook_dick['workbook'].sheetnames:
        sheet_for_pars = workbook_dick['workbook'].create_sheet(name_team)
        sheet_for_pars.cell(row=1, column=1).value = 'Вид спорта'
        sheet_for_pars.cell(row=1, column=2).value = 'Страна'
        sheet_for_pars.cell(row=1, column=3).value = 'Лига'
        sheet_for_pars.cell(row=1, column=4).value = 'Команда'
        sheet_for_pars.cell(row=1, column=5).value = 'Начало Матча'
        sheet_for_pars.cell(row=1, column=6).value = 'День'
        sheet_for_pars.cell(row=1, column=7).value = 'Месяц'
        sheet_for_pars.cell(row=1, column=8).value = 'Год'
        sheet_for_pars.cell(row=1, column=9).value = 'День Недели'
        sheet_for_pars.cell(row=1, column=10).value = 'Начало: Час'
        sheet_for_pars.cell(row=1, column=11).value = 'Начало: Минута'
        sheet_for_pars.cell(row=1, column=12).value = 'Хозяева A - На выезде H '
        sheet_for_pars.cell(row=1, column=13).value = 'Встреча команд не отформатирована'
        sheet_for_pars.cell(row=1, column=14).value = 'Команда 1'
        sheet_for_pars.cell(row=1, column=15).value = 'Команда 2'
        sheet_for_pars.cell(row=1, column=16).value = 'Счёт не отформатировано'
        sheet_for_pars.cell(row=1, column=17).value = 'Счёт: 1'
        sheet_for_pars.cell(row=1, column=18).value = 'Счёт: 2'
        sheet_for_pars.cell(row=1, column=19).value = 'Коэффициент: 1'
        sheet_for_pars.cell(row=1, column=20).value = 'Коэффициент: X'
        sheet_for_pars.cell(row=1, column=21).value = 'Коэффициент: 2'
        sheet_for_pars.cell(row=1, column=22).value = 'User Predictions: 1'
        sheet_for_pars.cell(row=1, column=23).value = 'User Predictions: X'
        sheet_for_pars.cell(row=1, column=24).value = 'User Predictions: 2'
        sheet_for_pars.cell(row=1, column=25).value = 'User Predictions: Проголосовало'
        sheet_for_pars.cell(row=1, column=26).value = 'Ссылка на игру'
        sheet_for_pars.cell(row=1, column=27).value = 'Позиция в архиве'
        sheet_for_pars.cell(row=1, column=28).value = 'Комментарий'


def save_in_xlsx(workbook_dick, link_match, name_team, info_to_math):
    """
    Сохранение информации в файл
    :param workbook_dick:
    :param link_match:
    :param name_team:
    :param info_to_math:
    :return:
    """
    sheet_for_pars = workbook_dick['workbook'][name_team]
    position = workbook_dick['workbook'][name_team].max_row
    sheet_for_pars.cell(row=position + 1, column=1).value = info_to_math['sport']
    sheet_for_pars.cell(row=position + 1, column=2).value = info_to_math['country']
    sheet_for_pars.cell(row=position + 1, column=3).value = info_to_math['liga']
    sheet_for_pars.cell(row=position + 1, column=4).value = info_to_math['name_team']
    sheet_for_pars.cell(row=position + 1, column=5).value = info_to_math['date_match_no_formatted']
    sheet_for_pars.cell(row=position + 1, column=6).value = info_to_math['day']
    sheet_for_pars.cell(row=position + 1, column=7).value = info_to_math['mouth']
    sheet_for_pars.cell(row=position + 1, column=8).value = info_to_math['years']
    sheet_for_pars.cell(row=position + 1, column=9).value = info_to_math['weekend']
    sheet_for_pars.cell(row=position + 1, column=10).value = info_to_math['hours']
    sheet_for_pars.cell(row=position + 1, column=11).value = info_to_math['minutes']
    sheet_for_pars.cell(row=position + 1, column=13).value = info_to_math['teams_not_formatted']
    sheet_for_pars.cell(row=position + 1, column=14).value = info_to_math['first_command']
    sheet_for_pars.cell(row=position + 1, column=15).value = info_to_math['second_command']
    sheet_for_pars.cell(row=position + 1, column=16).value = info_to_math['account_not_formatted']
    sheet_for_pars.cell(row=position + 1, column=19).value = info_to_math['coefficient_1']
    sheet_for_pars.cell(row=position + 1, column=20).value = info_to_math['coefficient_x']
    sheet_for_pars.cell(row=position + 1, column=21).value = info_to_math['coefficient_2']
    sheet_for_pars.cell(row=position + 1, column=22).value = info_to_math['user_predictions'][0]
    sheet_for_pars.cell(row=position + 1, column=23).value = info_to_math['user_predictions'][1]
    sheet_for_pars.cell(row=position + 1, column=24).value = info_to_math['user_predictions'][2]
    sheet_for_pars.cell(row=position + 1, column=25).value = info_to_math['amount_people_result']
    sheet_for_pars.cell(row=position + 1, column=26).value = link_match
    sheet_for_pars.cell(row=position + 1, column=27).value = position
    sheet_for_pars.cell(row=position + 1, column=28).value = info_to_math['comment']
    workbook_dick['workbook'].save(workbook_dick['file_name'])


def fill_up_all_data(workbook):
    """
    Функция собирает все матчи из выделенных команд и добавляет их в ALLDATA
    :param workbook: объект workbook
    :return: None
    """
    workbook['ALLDATA'].delete_rows(2, workbook['ALLDATA'].max_row)
    rows = workbook['CONTROL'].max_row
    flag = [cell.value for cell in workbook['CONTROL']['D'][1:] if cell.value is not None]
    for row in range(2, rows + 1):
        if len(flag) <= 1:
            break
        name_team = workbook['CONTROL'].cell(row=row, column=2).value
        if workbook['CONTROL'].cell(row=row, column=4).value and name_team in workbook.sheetnames:
            for i, info in enumerate(
                    workbook[name_team].iter_rows(max_col=28, max_row=workbook[name_team].max_row, values_only=True)):
                if i != 0:
                    workbook['ALLDATA'].append(info)


def transferring_functions_to_xl(workbook):
    """
    Функция вставляет формулы написанные пользователем
    :param workbook: объект workbook
    :return: None
    """
    rows = workbook['CONTROL'].max_row
    for row in range(2, rows + 1):
        if workbook['CONTROL'].cell(row=row, column=6).value:
            cells_xl = workbook['CONTROL'].cell(row=row, column=6).value.replace(' ', '')
            formula_xl = workbook['CONTROL'].cell(row=row, column=7).value
            if formula_xl is None:
                formula_xl = ''
            else:
                formula_xl = formula_xl.replace(' ', '').replace(';', ',')
            for name_list_xl in workbook.sheetnames:
                if name_list_xl == 'CONTROL':
                    continue
                rows_list = workbook[name_list_xl].max_row
                for row_list in range(2, rows_list + 1):
                    if formula_xl == '':
                        workbook[name_list_xl][f'{cells_xl}{row_list}'] = ''
                    else:
                        workbook[name_list_xl][f'{cells_xl}{row_list}'] = f'={formula_xl.replace("@", str(row_list))}'


def main(workbook_dick, links_to_team):
    options_firefox = webdriver.FirefoxOptions()
    options_firefox.add_argument(
        "user-agent=Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:84.0) Gecko/20100101 Firefox/84.0")
    options_firefox.add_argument("--disable-blink-features=AutomationControlled")
    options_firefox.add_argument("--headless")
    options_firefox.set_preference('permissions.default.image', 2)
    with webdriver.Firefox(options=options_firefox) as driver:  # executable_path='geckodriver',

    # options_chrome = webdriver.ChromeOptions()
    # options_chrome.add_argument("--disable-notifications")
    # options_chrome.add_argument(
    #     "user-agent=Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:84.0) Gecko/20100101 Firefox/84.0")
    # options_chrome.add_argument("--disable-popup-blocking")
    # options_chrome.add_argument('--headless')
    # options_chrome.add_argument('--disable-gpu')
    # with webdriver.Chrome(executable_path='chromedriver', options=options_chrome) as driver:
        login_acc(driver)
        """ Парсим игры команды"""
        for link_to_team in links_to_team:
            links_to_matches, name_team = parses_links_to_matches(driver, workbook_dick, link_to_team)
            """Удаляет последнии два матча из файла и добавляет ссылки к links_to_matches"""
            if workbook_dick['workbook'][name_team].max_row > 3:
                wb_name_team = workbook_dick['workbook'][name_team]
                links_to_matches.insert(0, wb_name_team.cell(row=wb_name_team.max_row - 1, column=26).value)
                links_to_matches.insert(0, wb_name_team.cell(row=wb_name_team.max_row, column=26).value)
                wb_name_team.delete_rows(wb_name_team.max_row - 1, 2)

            for a in workbook_dick['workbook']['CONTROL']['A']:
                if a.value == link_to_team:
                    workbook_dick['workbook']['CONTROL'].cell(row=a.row, column=a.column + 1).value = name_team
                """парсим матчи которых нет в файле"""
            for count, link_match in enumerate(links_to_matches[::-1]):
                count_break = 0
                while True:
                    try:
                        print(f'Match recording {count + 1} - {len(links_to_matches)} {link_match}')
                        info_to_math = parses_to_match(driver, link_match, name_team)
                        save_in_xlsx(workbook_dick, link_match, name_team, info_to_math)
                        break
                    except Exception:
                        print(f'Попыток сделано {count_break} из 5')
                        if count_break >= 5:
                            info_to_math = {
                                'name_team': name_team,
                                'sport': None,
                                'country': None,
                                'liga': None,
                                'date_match_no_formatted': None,
                                'day': None,
                                'mouth': None,
                                'years': None,
                                'weekend': None,
                                'hours': None,
                                'minutes': None,
                                'teams_not_formatted': None,
                                'first_command': None,
                                'second_command': None,
                                'account_not_formatted': None,
                                'first_goal': None,
                                'second_goal': None,
                                'coefficient_1': None,
                                'coefficient_x': None,
                                'coefficient_2': None,
                                'user_predictions': [None, None, None],
                                'amount_people_result': None,
                                'comment': 'Произошла ошибка при сборке информации',
                            }
                            save_in_xlsx(workbook_dick, link_match, name_team, info_to_math)
                            break
                        else:
                            time.sleep(5)
                    count_break += 1
            """парсим ещё не сыгранные матчи текущей команды матчи"""
            links_to_next_matches, _ = parses_links_to_matches(driver, workbook_dick,
                                                               link_to_team.replace('/results', ''))
            print(f'{len(links_to_next_matches)} matches will be added to NEXT MATCHES')

            for link_next_match in links_to_next_matches[::-1]:
                print(link_next_match)
                info_to_next_math = parses_to_match(driver, link_next_match, name_team)
                if info_to_next_math is not None:
                    save_in_xlsx(workbook_dick, link_next_match, 'NEXT MATCHES', info_to_next_math)
            workbook_dick['workbook'].save(workbook_dick['file_name'])

        """Наполняем ALLDATA матчами выбранных команд"""
        fill_up_all_data(workbook_dick['workbook'])
        workbook_dick['workbook'].save(workbook_dick['file_name'])
        """Вставляем формулы"""
        transferring_functions_to_xl(workbook_dick['workbook'])
        workbook_dick['workbook'].save(workbook_dick['file_name'])


if __name__ == '__main__':
    freeze_support()
    time_start = datetime.datetime.now()
    try:
        processes = []
        """Проходим циклом по всем .xlsx файла в папке и получаем ссылки на команды для парсинга"""
        for workbook_dick in open_xlsx_file():
            if 'CONTROL' in workbook_dick['workbook'].sheetnames:
                links_to_team = get_links_to_parse(workbook_dick['workbook']['CONTROL'])
                print(f'File in progress: {workbook_dick["file_name"]}')
            else:
                continue
            if 'NEXT MATCHES' not in workbook_dick['workbook'].sheetnames:
                create_new_sheet_if_not(workbook_dick, 'NEXT MATCHES')
            if 'ALLDATA' not in workbook_dick['workbook'].sheetnames:
                create_new_sheet_if_not(workbook_dick, 'ALLDATA')
            """очистка листа NEXT MATCHES"""
            workbook_dick['workbook']['NEXT MATCHES'].delete_rows(2, workbook_dick['workbook']['NEXT MATCHES'].max_row)
            """Создание процессов"""
            process = Process(target=main, args=(workbook_dick, links_to_team))
            processes.append(process)
            process.start()
        for process in processes:
            process.join()
        time_over = datetime.datetime.now()
        print(time_over - time_start)
        input('Данные успешно записаны!\nНажмите ENTER для закрытие программы')
    except Exception as log:
        time_error = datetime.datetime.now().strftime('%d.%m.%Y %H;%M;%S')
        reshetka = '###' * 15
        with open('LOG.txt', 'a', encoding='utf-8') as file:
            file.write(f'Ошибка возникла в {time_error}\n{reshetka}\n{log}')
        print(f'Ошибка возникла в {time_error}\n{reshetka}\n{log}')
        input('Произошла критическая ошибка!!!!!!!!!!!!!!!!!\nНажмите ENTER для закрытие программы')
